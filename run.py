from tokenize import Double
from zipfile import ZipFile
import os
import requests
from lxml import etree
from openpyxl import load_workbook
from openpyxl.styles import numbers

import datetime

# Téléchargement
print("Téléchargement du fichier")
remote_url = 'https://donnees.roulez-eco.fr/opendata/instantane'


prox = {
   'http': 'http://a430885:Bleu0033@138.21.17.49:3128',
   'https': 'http://a430885:Bleu0033@138.21.17.49:3128',
}



local_file = 'D:/DEV/carbu/PrixCarburants_instantane.zip'

data = requests.get(remote_url, proxies=prox)

with open(local_file, 'wb')as file:
    file.write(data.content)
  

# Décompression
with ZipFile(local_file, 'r') as zip: 
    zip.printdir() 
  
    print('extraction...') 
    zip.extractall('D:/DEV/carbu/') 
    print('Terminé!')


# Suppression
print("Suppression du fichier")
os.remove(local_file)
print("Fichier supprimé")


tree = etree.parse("D:/DEV/carbu/PrixCarburants_instantane.xml")

test = tree.xpath("/pdv_liste/pdv")

prix_diesel = 0
prix_sp95 = 0

for pdv in tree.xpath("/pdv_liste/pdv"):
    
    test_id = pdv.get("id")

    if test_id == "72230007":
        print(pdv[0]) #adresse
        print(pdv[1]) #ville

        for i in range(len(pdv)):
            

            if pdv[i].get("nom") == "Gazole" :
                print ("prix gazole : " + pdv[i].get("valeur"))
                prix_diesel = pdv[i].get("valeur")
                prix_num_diesel = prix_diesel
                prix_diesel = prix_diesel.replace(".", ",")

            if pdv[i].get("nom") == "SP95" :
                print ("prix SP95 : " + pdv[i].get("valeur"))
                prix_sp95 = pdv[i].get("valeur")
                prix_num_95 = prix_sp95
                prix_sp95 = prix_sp95.replace(".", ",")

# Suppression
print("Suppression du fichier")
os.remove("D:/DEV/carbu/PrixCarburants_instantane.xml")
print("Fichier supprimé")

# Remplissage fichier excel
workbook = load_workbook("D:/DEV/carbu/velotaf.xlsx", read_only=False)
sheet_vt = workbook["velotaf"]

day_date = datetime.date.today()
str_day_date = str(day_date.day) + "/" + str(day_date.month) + "/" + str(day_date.year)
print(day_date)

try:
    prix_num_95
except NameError:
    prix_num_95 = "0"

try:
    prix_num_diesel
except NameError:
    prix_num_diesel = "0"

for r in range(3, 10000):
    if str(sheet_vt.cell(row=r, column=2).value.date()) == str(day_date):
        print(str(r))
        sheet_vt.cell(row=r, column=4).number_format = '0.000'
        sheet_vt.cell(row=r, column=4).value = float(prix_num_95)

        sheet_vt.cell(row=r, column=5).number_format = '0.000'
        sheet_vt.cell(row=r, column=5).value = float(prix_num_diesel)



workbook.save("D:/DEV/carbu/velotaf.xlsx")